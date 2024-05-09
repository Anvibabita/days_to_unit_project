# using openpyxl library for working with the xls file
import openpyxl


inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

"""1) List each company with respective product count"""
# products_per_supplier = {}

# for product_row in range(2, product_list.max_row + 1): #product_row = 2 to 75
#     supplier_name = product_list.cell(product_row, 4).value #supplier_name = AAA Company,BBB Company,CCC Company.....
    
#     # calculation number of products per supplier
#     if supplier_name in products_per_supplier:
#         current_num_products = products_per_supplier.get(supplier_name)
#         products_per_supplier[supplier_name] = current_num_products + 1
#     else:
#         products_per_supplier[supplier_name] = 1
        
# print(products_per_supplier)       

""" 2) List each company with respective total inventory value"""     
# products_per_supplier = {}
# total_value_per_supplier = {}

# for product_row in range(2, product_list.max_row + 1): 
#     supplier_name = product_list.cell(product_row, 4).value 
#     inventory = product_list.cell(product_row, 2).value
#     price = product_list.cell(product_row, 3).value
    
#     #1 calculation number of products per supplier
#     if supplier_name in products_per_supplier:
#         current_num_products = products_per_supplier.get(supplier_name)
#         products_per_supplier[supplier_name] = current_num_products + 1
#     else:
#         products_per_supplier[supplier_name] = 1
        
    
#     #2 calculation_total_value_of_inventory_per_supplier
#     if supplier_name in total_value_per_supplier:
#         current_total_value = total_value_per_supplier.get(supplier_name)
#         total_value_per_supplier[supplier_name] = current_total_value +inventory * price
    
#     else:    
#         total_value_per_supplier[supplier_name] =  inventory * price
    

# print(products_per_supplier)    
# print(total_value_per_supplier)    

""" 3) List products with inventory less than 10 """
# products_per_supplier = {}
# total_value_per_supplier = {}
# product_under_10_inv = {}

# for product_row in range(2, product_list.max_row + 1): 
#     supplier_name = product_list.cell(product_row, 4).value 
#     inventory = product_list.cell(product_row, 2).value
#     price = product_list.cell(product_row, 3).value
#     product_num = product_list.cell(product_row, 1).value
    
#     #1 calculation number of products per supplier
#     if supplier_name in products_per_supplier:
#         current_num_products = products_per_supplier.get(supplier_name)
#         products_per_supplier[supplier_name] = current_num_products + 1
#     else:
#         products_per_supplier[supplier_name] = 1
        
    
#     #2 calculation total value of inventory per supplier
#     if supplier_name in total_value_per_supplier:
#         current_total_value = total_value_per_supplier.get(supplier_name)
#         total_value_per_supplier[supplier_name] = current_total_value +inventory * price
    
#     else:    
#         total_value_per_supplier[supplier_name] =  inventory * price
        
        
#     # logic products with inventory less than 10
#     if inventory < 10: 
#         product_under_10_inv[product_num] = inventory #{25: 7, 30: 6, 74: 2} but if you are getting {25.0: 7.0, 30.0: 6.0, 74.0: 2.0} so change int(inventory)
        
    

# # print(products_per_supplier)    
# # print(total_value_per_supplier)    
# print(product_under_10_inv) #{25: 7, 30: 6, 74: 2}

""" 4) write to spreadsheet: calculate and write inventory value for each product into spreadsheet """
products_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1): 
    supplier_name = product_list.cell(product_row, 4).value 
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    
    #1 calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1
        
    
    #2 calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value +inventory * price
    
    else:    
        total_value_per_supplier[supplier_name] =  inventory * price
        
        
    #3 logic products with inventory less than 10
    if inventory < 10: 
        product_under_10_inv[product_num] = inventory #{25: 7, 30: 6, 74: 2} but if you are getting {25.0: 7.0, 30.0: 6.0, 74.0: 2.0} so change int(inventory)
        
    
    #4 add value for total inventory price
    inventory_price.value = inventory * price


print(products_per_supplier)    
print(total_value_per_supplier)    
print(product_under_10_inv) #{25: 7, 30: 6, 74: 2}

inv_file.save("inventory_with_total_value.xlsx")